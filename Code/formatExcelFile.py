''' formatExcelFile.py '''
from openpyxl.styles import Font, Alignment
alignment1 = Alignment(wrapText=True,
                       horizontal='left',
                       vertical='center')
alignment2 = Alignment(horizontal='center',
                       vertical='center')
ft = Font(name='Tahoma', bold=True, size="11")


def formatExcelFile(ws):
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.cell(10, 2).value = "Amount"
    ws.cell(10, 3).value = "Description"
    ws = formatCommon(ws, 10)
    ws.cell(10, 4).value = "Category"
    for row in range(1, ws.max_row + 1):
        ws.cell(row, 2).number_format = "0.00"
    return ws


def formatCommon(ws, row):
    ws.cell(row, 1).value = "Date"
    ws.print_options.gridLines = True
    ws.cell(row, 2).number_format = "0.00"
    for col in range(1, 5):
        ws.cell(row, col).font = ft
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 18
        ws.cell(row, 1).number_format = "DD MM YY"
        for col in range(1, 5):
            ws.cell(row, col).alignment = alignment1
            # ws.cell(row, col).font = ft
            ws.cell(row, 1).alignment = alignment2
    return ws


def formatCategoryWs(ws, theList):
    ws.cell(1, 3).value = "Amount"
    ws.cell(1, 2).value = "Description"
    ws.freeze_panes = 'E2'
    ws.auto_filter.ref = "A1:C" + str(ws.max_row)
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws = formatCommon(ws, 1)
    return ws