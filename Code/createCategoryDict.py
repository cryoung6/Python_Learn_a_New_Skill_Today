''' createCategoryDict.py

    Create a dictionary of the categories of expenses
    from the Excel file Categories of Expenses.xlsx
    the phrase in Column 1 will be the key
    expCategories_Dict[phrase] = category
'''


def createCategoryDict():
    import os
    from openpyxl import load_workbook

    codeDir = os.getcwd()
    source = codeDir + "/source/"
    filename = source + "Categories of Expenses.xlsx"

    wbCategories = load_workbook(filename, data_only=True)
    wsCategories = wbCategories["Sheet1"]

    expCategories_Dict = {}
    phrase, category = '', ''

    for row in range(2, wsCategories.max_row + 1):
        phrase = wsCategories.cell(row, 1).value
        phrase = phrase.upper()
        category = wsCategories.cell(row, 2).value.upper()
        expCategories_Dict[phrase] = category

    return expCategories_Dict
