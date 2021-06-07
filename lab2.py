import string
from openpyxl import Workbook
import os
import sys
currentWorkingDir = os.getcwd()
codeDir = currentWorkingDir + '/code'
# On Mac add /code folder to path. On a Windows PC it would be \\code
sys.path.insert(0, codeDir)


# these modules are in the /code subdirectory
from createCategoryDict import createCategoryDict
from getCategory import getCategory
from getFilename import getFilename
from writeSumTotalsToFile import writeSumTotalsToFile
from formatExcelFile import formatExcelFile
from categoryWorksheets import categoryWorksheets

'''  Keep a running total of the category expenses. These are float
     variables.
         sumAdv
         sumDepr
         sumOther
         sumProf

     In the Excel output file these will be the first few lines.
         Row 1 Column 2 sumAdv
         Row 1 Column 3 "Advertising Total"
         Row 2 Column 2 sumProf
         Row 2 Column 3 'Legal & Professional Total'
         Row 3 Column 2 sumOther
         Row 3 Column 3 'Other Publishing Expenses Total'
         Row 4 Column 2 sumDepr
         Row 4 Column 3 'Depreciation & Section 179 Total'
'''


def sumTotals(category, thisAmt, tempList):
    global sumAdv, sumOther, sumProf, sumOffice
    global sumSupplies, sumMisc, sumAssets
    global advList, otherList, profList, miscList
    global officeList, assetsList, suppliesList

    if category == 'ADVERTISING':
        sumAdv += float(thisAmt)
        advList.append(tempList)
    elif category == 'LEGAL & PROFESSIONAL':
        sumProf += float(thisAmt)
        profList.append(tempList)
    elif category == 'OTHER PUBLISHING EXPENSES':
        sumOther += float(thisAmt)
        otherList.append(tempList)
    elif category == 'OFFICE EXPENSES':
        sumOffice += float(thisAmt)
        officeList.append(tempList)
    elif category == 'ASSETS':
        sumAssets += float(thisAmt)
        assetsList.append(tempList)
    elif category == 'SUPPLIES':
        sumSupplies += float(thisAmt)
        suppliesList.append(tempList)
    elif category == 'OTHER MISC EXPENSES':
        sumMisc += float(thisAmt)
        miscList.append(tempList)


sumAdv, sumOther, sumProf, sumMisc = 0.0, 0.0, 0.0, 0.0
sumOffice, sumAssets, sumSupplies = 0.0, 0.0, 0.0
advList,  otherList, profList, miscList = [], [], [], []
officeList, assetsList, suppliesList = [], [], []


expCategories_Dict = {}
expCategories_Dict = createCategoryDict()


row_out = 11
wb = Workbook()
ws = wb.create_sheet('Business_Expenses')
wb.remove(wb['Sheet'])


expenses_Dict = {}


numbers = string.digits


mypath1 = os.environ['HOME']
mypath2 = os.path.join(mypath1, 'BOOK_AUTHORING', 'Python_Lab2',
                       'Code_for_Lab2', 'Receipts')
files = sorted(os.listdir(mypath2))


# main body of the program
for i in range(len(files)):
    thisFileName = str(files[i])
    lenFileName = len(thisFileName)
    if thisFileName[0] in numbers:
        thisDate, thisDesc, thisCategory = '', '', ''
        tempList = []
        dollarSignIndex, end, thisAmt = 0, 0, 0.0

        # get date from first 6 characters of file name
        thisDate += thisFileName[:2]
        thisDate += r'/'
        thisDate += thisFileName[2:4]
        thisDate += r'/'
        thisDate += thisFileName[4:6]

        if '$' in thisFileName:
            dollarSignIndex = thisFileName.rindex('$') + 1
            end = thisFileName.rindex('.')
            thisAmt = thisFileName[dollarSignIndex:end]
        else:
            print('No $ Amount in Filename', thisFileName)

        thisDesc = thisFileName[7:dollarSignIndex - 1].rstrip()

        ws.cell(row_out, 1).value = thisDate
        ws.cell(row_out, 2).value = float(thisAmt)
        ws.cell(row_out, 3).value = thisDesc

        thisCategory = getCategory(thisFileName, expCategories_Dict)
        tempList = [thisDate, thisDesc, thisAmt, thisCategory]

        if not thisCategory and not thisAmt:
            print("Couldn't find category or amount for: ", thisFileName)
        else:
            sumTotals(thisCategory, thisAmt, tempList)
            ws.cell(row_out, 4).value = thisCategory
        row_out += 1

        expenses_Dict[thisFileName] = tempList


# end of "for loop" that began on line 96

# add each category's sumtotal to rows 1-5 in the "ws" worksheet
ws = writeSumTotalsToFile(ws, sumAdv, sumProf, sumOther, sumOffice,
                          sumAssets, sumSupplies, sumMisc)

# create the four worksheets
advList.sort()
otherList.sort()
profList.sort()
officeList.sort()
assetsList.sort()
suppliesList.sort()
miscList.sort()
wb = categoryWorksheets(wb, advList, otherList, profList, officeList,
                        assetsList, suppliesList, miscList)

ws = formatExcelFile(ws)

wb.save(getFilename())
wb.save('Business_Expenses.xlsx')


total_ReceiptsCategorized = ws.max_row - 9
total_Files = len(files)
msg1 = "There are " + str(total_Files) + " files and"
msg2 = str(total_ReceiptsCategorized) + " categorized receipts."
print(msg1, msg2)


























